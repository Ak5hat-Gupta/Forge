from __future__ import annotations

import csv
import io
from typing import Any

from app.services.inference import ColumnInference, infer_schema


def parse_csv(content: bytes) -> tuple[list[str], list[list[str | None]]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows_raw = list(reader)
    if not rows_raw:
        return [], []
    headers = [h.strip() for h in rows_raw[0]]
    data = rows_raw[1:]
    return headers, data


def parse_excel(content: bytes) -> tuple[list[str], list[list[str | None]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows_raw:
        return [], []
    headers = [str(h).strip() if h is not None else f"column_{i}" for i, h in enumerate(rows_raw[0])]
    data = [[str(cell) if cell is not None else None for cell in row] for row in rows_raw[1:]]
    return headers, data


def ingest_file(
    filename: str, content: bytes
) -> tuple[list[ColumnInference], list[dict[str, Any]]]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext in ("xlsx", "xls"):
        headers, data = parse_excel(content)
    else:
        headers, data = parse_csv(content)

    if not headers:
        return [], []

    schema = infer_schema(headers, data)
    slug_map = {i: col.slug for i, col in enumerate(schema)}

    rows: list[dict[str, Any]] = []
    for raw_row in data:
        row_data: dict[str, Any] = {}
        for i, col in enumerate(schema):
            val = raw_row[i] if i < len(raw_row) else None
            if val is not None and val.strip() == "":
                val = None
            row_data[col.slug] = _coerce(val, col.inferred_type)
        rows.append(row_data)

    return schema, rows


def _coerce(value: str | None, inferred_type: str) -> Any:
    if value is None:
        return None
    v = value.strip()
    if not v:
        return None

    try:
        if inferred_type == "boolean":
            return v.lower() in ("true", "yes", "1", "y", "t")
        if inferred_type == "integer":
            return int(v.replace(",", ""))
        if inferred_type == "float":
            return float(v.replace(",", ""))
        if inferred_type == "currency":
            return float(v.lstrip("$€£¥ ").replace(",", ""))
        return v
    except (ValueError, TypeError):
        return v
